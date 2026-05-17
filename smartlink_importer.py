#!/usr/bin/env python3
"""
SIJI Smartlink Import Pipeline
Import Smartlink XLSX exports into PostgreSQL siji_db.

XLSX Structure (Smartlink "Rekap Data Transaksi Reguler"):
  Rows 1-23: Summary/metadata
  Row 24: Header (No, No Nota, Customer, No Telp, Alamat, etc.)
  Row 25: Empty (sometimes)
  Row 26+: Data rows (merged cells for multi-item transactions)
  
  Cols 1-21: Transaction header
  Cols 22-28: Detail items (Tipe, Nama Layanan, Qty, Satuan, Harga, Jumlah, Keterangan)
"""

import os, sys, re, json, argparse
from datetime import datetime

script_dir = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, script_dir)

from product_mapping import classify_product
from address_normalizer import LocationNormalizer

MONTHS = {
    'jan':'01','feb':'02','mar':'03','apr':'04','mei':'05','may':'05',
    'jun':'06','jul':'07','agu':'08','aug':'08','ags':'08','agus':'08',
    'sep':'09','sept':'09','okt':'10','oct':'10','nov':'11',
    'nop':'11','des':'12','dec':'12'
}

def parse_indo_date(s):
    if not s or str(s).strip() in ('-','','None'): return None
    s = str(s).strip()
    if re.match(r'^\d{4}-\d{2}-\d{2}', s): return s[:10]
    m = re.match(r'(\d{1,2})\s+(\w+)\s+(\d{4})', s)
    if m:
        day, mon, year = m.groups()
        mn = MONTHS.get(mon.lower()[:4]) or MONTHS.get(mon.lower()[:3])
        if mn: return f"{year}-{mn}-{int(day):02d}"
    return None

def normalize_phone(phone):
    if not phone: return None
    p = re.sub(r'[^\d]', '', str(phone))
    if not p: return None
    if p.startswith('0'): p = '62' + p[1:]
    if not p.startswith('62'): p = '62' + p
    return p

def safe_float(v, default=0.0):
    try: return float(v) if v else default
    except: return default


class SmartlinkImporter:
    def __init__(self, pg_conn, normalizer=None):
        self.pg = pg_conn
        self.normalizer = normalizer
        self.stats = {'total':0,'parsed':0,'inserted':0,'updated':0,'duplicate':0,'failed':0,'skipped':0}
        self.errors = []

    def import_file(self, filepath, dry_run=False, normalize_addr=True):
        groups = self._parse_xlsx(filepath)
        self.stats['total'] = len(groups)
        print(f"Parsed {len(groups)} nota from {os.path.basename(filepath)}")
        
        if not groups:
            print("No data!"); return self.stats

        # Validate each group (header from first item)
        valid_groups = []
        for nota, items in groups.items():
            header = items[0]
            issues = self._validate(header)
            if issues:
                self.stats['failed'] += 1
                if len(self.errors) < 20:
                    self.errors.append(f"Nota {nota}: {'; '.join(issues)}")
            else:
                valid_groups.append((nota, items))

        self.stats['parsed'] = len(valid_groups)
        print(f"Valid: {len(valid_groups)}, Failed: {self.stats['failed']}")
        if self.errors:
            print("Errors (first 5):")
            for e in self.errors[:5]: print(f"  {e}")

        if dry_run:
            print("\n[DRY RUN] Preview first 3:")
            for nota, items in valid_groups[:3]:
                header = items[0]
                print(f"  {nota} | {header['tgl_terima']} | {header['customer_name']} | {len(items)} items | {header['total_tagihan']}")
            return self.stats

        cur = self.pg.cursor()
        for nota, items in valid_groups:
            try:
                self._upsert_group(cur, nota, items, normalize_addr)
            except Exception as e:
                self.stats['failed'] += 1
                self.errors.append(f"DB: {nota}: {e}")
                self.pg.rollback()

        # Log
        cur.execute("""INSERT INTO import_logs (filename, imported_by, total_rows, success_rows, failed_rows, duplicate_rows, quality_score)
            VALUES (%s,'script',%s,%s,%s,%s,%s)""",
            (os.path.basename(filepath), self.stats['total'],
             self.stats['inserted']+self.stats['updated'], self.stats['failed'],
             self.stats['duplicate'], self._quality()))
        self.pg.commit()
        self._summary()
        return self.stats

    def _parse_xlsx(self, filepath):
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=False)
        ws = wb.active

        # Unmerge all cells
        for mc in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(mc))

        # Find header row
        header_row = None
        for rn in range(1, 40):
            vals = [str(c.value or '') for c in ws[rn]]
            if any('No Nota' in v for v in vals):
                header_row = rn; break
        if not header_row:
            raise ValueError("Header not found (looking for 'No Nota')")
        print(f"Header at row {header_row}")

        # Map header names to column indices
        hdr = {}
        for i, cell in enumerate(ws[header_row]):
            if cell.value: hdr[str(cell.value).strip()] = i

        # Column mapping
        COL = {
            'No Nota':'no_nota', 'Customer':'customer_name', 'No Telp Customer':'customer_phone',
            'Alamat Customer':'customer_address', 'Progres Pengerjaan':'progress',
            'Outlet':'outlet', 'Tgl Terima':'tgl_terima_raw', 'Tgl Selesai':'tgl_selesai_raw',
            'Subtotal':'subtotal', 'Tambahan Express':'tambahan_express', 'Diskon':'diskon',
            'Pajak':'pajak', 'Biaya Service':'biaya_service', 'Total Tagihan':'total_tagihan',
            'Jenis':'jenis', 'Pembayaran':'pembayaran', 'Pengambilan':'pengambilan',
            'Tgl Pengambilan':'tgl_pengambilan_raw', 'Pembuat Nota':'kasir', 'Keterangan Nota':'catatan'
        }
        cidx = {}
        for hname, key in COL.items():
            if hname in hdr: cidx[key] = hdr[hname]

        # Parse rows and group by no_nota
        groups = {}  # no_nota -> list of item rows
        current_nota = None

        for rn in range(header_row + 1, ws.max_row + 1):
            cells = [c.value for c in ws[rn]]
            if not any(c for c in cells): continue

            # Detect nota row (col 1 has SJ...)
            no_nota = cells[cidx['no_nota']] if 'no_nota' in cidx and cidx['no_nota'] < len(cells) else None
            if no_nota and str(no_nota).startswith('SJ'):
                current_nota = str(no_nota).strip()
            elif no_nota:
                continue  # skip non-SJ rows
            elif current_nota is None:
                continue  # skip rows before first nota

            row = {'no_nota': current_nota}
            for key, idx in cidx.items():
                row[key] = cells[idx] if idx < len(cells) else None

            # Detail columns: 21=jenis/tipe, 22=nama, 23=jumlah, 24=satuan, 25=total, 26=jumlah_item, 27=keterangan
            row['detail_jenis'] = cells[21] if len(cells) > 21 else None
            row['nama_layanan'] = cells[22] if len(cells) > 22 else None
            row['jumlah'] = safe_float(cells[23] if len(cells) > 23 else None)
            row['satuan'] = cells[24] if len(cells) > 24 else None
            row['total_item'] = safe_float(cells[25] if len(cells) > 25 else None)
            row['jumlah_item'] = safe_float(cells[26] if len(cells) > 26 else None)
            row['keterangan_layanan'] = cells[27] if len(cells) > 27 else None

            # Parse dates (only on header row, copy to sub-items)
            if row.get('tgl_terima_raw'):
                row['tgl_terima'] = parse_indo_date(row.get('tgl_terima_raw'))
            if row.get('tgl_selesai_raw'):
                row['tgl_selesai'] = parse_indo_date(row.get('tgl_selesai_raw'))
            if row.get('tgl_pengambilan_raw'):
                row['tgl_pengambilan'] = parse_indo_date(row.get('tgl_pengambilan_raw'))

            # Normalize amounts (only on header row)
            if row.get('subtotal') is not None or row.get('total_tagihan'):
                for f in ['subtotal','tambahan_express','diskon','pajak','biaya_service','total_tagihan']:
                    row[f] = safe_float(row.get(f))

            # Phone (only on header row)
            if row.get('customer_phone'):
                row['customer_phone'] = normalize_phone(row.get('customer_phone'))

            if current_nota not in groups:
                groups[current_nota] = []
            groups[current_nota].append(row)

        wb.close()
        return groups

    def _validate(self, row):
        issues = []
        if not row.get('no_nota'): issues.append("Missing no_nota")
        if not row.get('tgl_terima'): issues.append(f"Bad date: {row.get('tgl_terima_raw')}")
        # Multi-item transactions: sub-items may have total=0 but nama_layanan set
        # These are valid - they share the parent's total. Skip amount check for sub-items.
        if row.get('total_tagihan', 0) <= 0 and row.get('subtotal', 0) <= 0:
            # Only fail if this is clearly a bad row (no product info either)
            if not row.get('nama_layanan'):
                issues.append("Zero amount and no product")
            else:
                self.stats['skipped'] += 1
                issues.append("SKIP_SUBITEM")  # Will be filtered, not counted as error
        return issues

    def _upsert_group(self, cur, no_nota, items, normalize_addr):
        header = items[0]
        no_nota = str(no_nota).strip()
        
        # Check duplicate
        cur.execute("SELECT id FROM transactions WHERE no_nota = %s", (no_nota,))
        existing = cur.fetchone()
        
        # Classify first item
        first_name = header.get('nama_layanan')
        cls = classify_product(first_name) if first_name else {}
        
        # Normalize address
        addr_norm = None
        addr_conf = None
        if normalize_addr and self.normalizer and header.get('customer_address'):
            nr = self.normalizer.normalize(header['customer_address'])
            addr_norm = nr['normalized']
            addr_conf = nr['confidence']

        # Match customer
        cust_id = None
        if header.get('customer_phone'):
            cur.execute("SELECT id FROM customers WHERE nomor_telpon = %s", (header['customer_phone'],))
            r = cur.fetchone()
            if r: cust_id = r[0]
        if not cust_id and header.get('customer_name'):
            cur.execute("SELECT id FROM customers WHERE LOWER(nama) = %s", (header['customer_name'].strip().lower(),))
            r = cur.fetchone()
            if r: cust_id = r[0]

        if existing:
            # Update header
            cur.execute("""UPDATE transactions SET
                customer_id=%s, customer_name=%s, customer_phone=%s, customer_address=%s,
                progress_status=%s, outlet=%s, tgl_terima_orig=%s, date_of_transaction=%s,
                tgl_selesai=%s, tgl_pengambilan=%s,
                subtotal=%s, tambahan_express=%s, diskon=%s, pajak=%s, total_tagihan=%s,
                jenis=%s, pembayaran=%s, pengambilan=%s, pembuat_nota=%s, keterangan_nota=%s,
                nama_layanan=%s, group_layanan=%s, jumlah=%s, satuan=%s, total_item=%s, jumlah_item=%s,
                keterangan_layanan=%s, customer_address_normalized=%s, address_confidence=%s,
                import_file=%s, imported_at=NOW()
                WHERE no_nota=%s""",
                (cust_id, header.get('customer_name'), header.get('customer_phone'), header.get('customer_address'),
                 header.get('progress'), header.get('outlet','SIJI Bintaro'), header.get('tgl_terima_raw'), header.get('tgl_terima'),
                 header.get('tgl_selesai'), header.get('tgl_pengambilan'),
                 header.get('subtotal',0), header.get('tambahan_express',0), header.get('diskon',0), header.get('pajak',0),
                 header.get('total_tagihan',0),
                 header.get('jenis','Reguler'), header.get('pembayaran'), header.get('pengambilan'),
                 header.get('kasir'), header.get('catatan'),
                 first_name, cls.get('kategori'),
                 items[0].get('jumlah'), items[0].get('satuan'), items[0].get('total_item'), items[0].get('jumlah_item'),
                 items[0].get('keterangan_layanan'), addr_norm, addr_conf,
                 os.path.basename(header.get('_filepath', 'smartlink_import')),
                 no_nota))
            # Delete old details
            cur.execute("DELETE FROM transaction_details WHERE no_nota = %s", (no_nota,))
            self.stats['updated'] += 1
        else:
            # Insert header
            cur.execute("""INSERT INTO transactions 
                (no_nota, customer_id, customer_name, customer_phone, customer_address,
                 progress_status, outlet, tgl_terima_orig, date_of_transaction, tgl_selesai, tgl_pengambilan,
                 subtotal, tambahan_express, diskon, pajak, total_tagihan,
                 jenis, pembayaran, pengambilan, pembuat_nota, keterangan_nota,
                 nama_layanan, group_layanan, jumlah, satuan, total_item, jumlah_item, keterangan_layanan,
                 import_file, imported_at, customer_address_normalized, address_confidence)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW(),%s,%s)""",
                (no_nota, cust_id, header.get('customer_name'), header.get('customer_phone'),
                 header.get('customer_address'), header.get('progress'), header.get('outlet','SIJI Bintaro'),
                 header.get('tgl_terima_raw'), header.get('tgl_terima'), header.get('tgl_selesai'), header.get('tgl_pengambilan'),
                 header.get('subtotal',0), header.get('tambahan_express',0), header.get('diskon',0), header.get('pajak',0),
                 header.get('total_tagihan',0),
                 header.get('jenis','Reguler'), header.get('pembayaran'), header.get('pengambilan'),
                 header.get('kasir'), header.get('catatan'),
                 first_name, cls.get('kategori'),
                 items[0].get('jumlah'), items[0].get('satuan'), items[0].get('total_item'), items[0].get('jumlah_item'),
                 items[0].get('keterangan_layanan'),
                 os.path.basename(header.get('_filepath', 'smartlink_import')),
                 addr_norm, addr_conf))
            self.stats['inserted'] += 1

        # Insert ALL items into transaction_details
        for line_num, item in enumerate(items, 1):
            item_name = item.get('nama_layanan')
            item_cls = classify_product(item_name) if item_name else {}
            cur.execute("""
                INSERT INTO transaction_details
                (no_nota, line_number, jenis, nama_layanan, group_layanan,
                 jumlah, satuan, total_item, jumlah_item, keterangan)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                no_nota, line_num,
                item.get('detail_jenis'), item_name, item_cls.get('kategori'),
                item.get('jumlah'), item.get('satuan'), item.get('total_item'), item.get('jumlah_item'),
                item.get('keterangan_layanan')
            ))

    def _quality(self):
        total = self.stats['total'] or 1
        return round((self.stats['inserted'] + self.stats['updated']) / total * 100, 1)

    def _summary(self):
        print(f"\n{'='*50}")
        print(f"Import Summary")
        print(f"{'='*50}")
        print(f"  Total rows:    {self.stats['total']}")
        print(f"  Parsed:        {self.stats['parsed']}")
        print(f"  Inserted:      {self.stats['inserted']}")
        print(f"  Updated:       {self.stats['updated']}")
        print(f"  Failed:        {self.stats['failed']}")
        print(f"  Quality:       {self._quality()}%")
        print(f"{'='*50}")


def main():
    parser = argparse.ArgumentParser(description='SIJI Smartlink Importer')
    parser.add_argument('action', choices=['import', 'validate'])
    parser.add_argument('--file', '-f', required=True, help='Path to XLSX/CSV file')
    parser.add_argument('--dry-run', action='store_true', help='Preview without importing')
    parser.add_argument('--no-normalize', action='store_true', help='Skip address normalization')
    parser.add_argument('--pg-host', default='localhost')
    parser.add_argument('--pg-port', type=int, default=5432)
    parser.add_argument('--pg-db', default='siji_db')
    parser.add_argument('--pg-user', default='siji')
    parser.add_argument('--pg-pass', default='siji2026db')
    args = parser.parse_args()

    if not os.path.exists(args.file):
        print(f"File not found: {args.file}"); sys.exit(1)

    import psycopg2
    pg = psycopg2.connect(host=args.pg_host, port=args.pg_port, dbname=args.pg_db, 
                          user=args.pg_user, password=args.pg_pass)
    
    # Init normalizer
    ref_path = os.path.join(script_dir, 'location_references.json')
    if not os.path.exists(ref_path):
        ref_path = os.path.join(script_dir, '..', 'data', 'location_references.json')
    normalizer = LocationNormalizer(reference_path=ref_path, pg_conn=pg) if not args.no_normalize else None

    importer = SmartlinkImporter(pg, normalizer)

    if args.action == 'validate':
        importer.import_file(args.file, dry_run=True)
    else:
        importer.import_file(args.file, dry_run=args.dry_run, normalize_addr=not args.no_normalize)

    pg.close()

if __name__ == '__main__':
    main()

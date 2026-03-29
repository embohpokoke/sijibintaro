from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _rp(amount: Any) -> str:
    try:
        value = int(amount or 0)
    except (TypeError, ValueError):
        value = 0
    return f"Rp {value:,.0f}".replace(",", ".")


def build_accounting_excel(report: dict[str, Any]) -> io.BytesIO:
    workbook = Workbook()
    summary_ws = workbook.active
    summary_ws.title = "Summary"
    pengeluaran_ws = workbook.create_sheet("Pengeluaran")
    pemasukan_ws = workbook.create_sheet("Pemasukan")
    mutasi_ws = workbook.create_sheet("Mutasi")

    title_fill = PatternFill("solid", fgColor="1F4E78")
    header_fill = PatternFill("solid", fgColor="D9E2F3")
    title_font = Font(color="FFFFFF", bold=True, size=14)
    header_font = Font(bold=True)

    periode = report["summary"]["periode_label"]
    summary_rows = [
        ["Laporan Keuangan SIJI Bintaro", ""],
        ["Periode", periode],
        ["Total Pemasukan", report["summary"]["total_pemasukan"]],
        ["Total Pengeluaran", report["summary"]["total_pengeluaran"]],
        ["Net Profit", report["summary"]["net_profit"]],
        ["Mutasi Belum Ditautkan", report["summary"]["unlinked_mutasi"]],
    ]

    for row_idx, row in enumerate(summary_rows, start=1):
        for col_idx, value in enumerate(row, start=1):
            cell = summary_ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center")
            if row_idx == 1:
                cell.fill = title_fill
                cell.font = title_font
            elif row_idx == 2:
                cell.font = header_font
            elif row_idx >= 3 and col_idx == 2 and row_idx <= 5:
                cell.number_format = '#,##0'
    summary_ws.merge_cells("A1:B1")

    summary_ws["A8"] = "Breakdown Pengeluaran"
    summary_ws["A8"].font = header_font
    summary_ws["A8"].fill = header_fill
    summary_ws["A9"] = "Kategori"
    summary_ws["B9"] = "Nominal"
    summary_ws["A9"].font = summary_ws["B9"].font = header_font
    summary_ws["A9"].fill = summary_ws["B9"].fill = header_fill
    for idx, row in enumerate(report["summary"]["breakdown_pengeluaran"], start=10):
        summary_ws.cell(row=idx, column=1, value=row["label"])
        nominal_cell = summary_ws.cell(row=idx, column=2, value=row["nominal"])
        nominal_cell.number_format = '#,##0'

    summary_ws["D8"] = "Breakdown Pemasukan"
    summary_ws["D8"].font = header_font
    summary_ws["D8"].fill = header_fill
    summary_ws["D9"] = "Kategori"
    summary_ws["E9"] = "Nominal"
    summary_ws["D9"].font = summary_ws["E9"].font = header_font
    summary_ws["D9"].fill = summary_ws["E9"].fill = header_fill
    for idx, row in enumerate(report["summary"]["breakdown_pemasukan"], start=10):
        summary_ws.cell(row=idx, column=4, value=row["label"])
        nominal_cell = summary_ws.cell(row=idx, column=5, value=row["nominal"])
        nominal_cell.number_format = '#,##0'

    _write_table(
        pengeluaran_ws,
        ["Tanggal", "Deskripsi", "Kategori", "Supplier", "Metode", "Dicatat Oleh", "Nominal", "Referensi", "Catatan"],
        [
            [
                row["tanggal"],
                row["deskripsi"],
                row["kategori_label"],
                row["supplier_nama"],
                row["metode_bayar"],
                row["dicatat_oleh"],
                row["nominal"],
                row["no_referensi"],
                row["catatan"],
            ]
            for row in report["pengeluaran"]
        ],
        title_fill,
        title_font,
        header_fill,
        header_font,
        money_columns={7},
    )

    _write_table(
        pemasukan_ws,
        ["Tanggal", "Customer", "Layanan", "Kategori", "Metode", "Status", "Nominal", "Dicatat Oleh", "Catatan"],
        [
            [
                row["tanggal"],
                row["nama_customer"],
                row["layanan"],
                row["kategori"],
                row["metode_bayar"],
                row["status"],
                row["nominal"],
                row["dicatat_oleh"],
                row["catatan"],
            ]
            for row in report["pemasukan"]
        ],
        title_fill,
        title_font,
        header_fill,
        header_font,
        money_columns={7},
    )

    _write_table(
        mutasi_ws,
        ["Tanggal", "Tipe", "Nominal", "No Urut", "Keterangan", "Penerima", "No Rek", "Status Link"],
        [
            [
                row["tanggal"],
                row["tipe"],
                row["nominal"],
                row["no_urut"],
                row["keterangan"],
                row["penerima"],
                row["no_rek_penerima"],
                row["status_link"],
            ]
            for row in report["mutasi"]
        ],
        title_fill,
        title_font,
        header_fill,
        header_font,
        money_columns={3},
    )

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def _write_table(
    worksheet,
    headers: list[str],
    rows: list[list[Any]],
    title_fill: PatternFill,
    title_font: Font,
    header_fill: PatternFill,
    header_font: Font,
    money_columns: set[int] | None = None,
) -> None:
    money_columns = money_columns or set()
    worksheet["A1"] = worksheet.title
    worksheet["A1"].fill = title_fill
    worksheet["A1"].font = title_font
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    for idx, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=2, column=idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, start=3):
        for col_idx, value in enumerate(row, start=1):
            cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
            if col_idx in money_columns:
                cell.number_format = '#,##0'
            if isinstance(value, str) and len(value) > 30:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    for col_idx in range(1, len(headers) + 1):
        column = get_column_letter(col_idx)
        max_length = max(
            len(str(worksheet.cell(row=row_idx, column=col_idx).value or ""))
            for row_idx in range(1, max(worksheet.max_row, 1) + 1)
        )
        worksheet.column_dimensions[column].width = min(max(max_length + 2, 12), 32)


def build_accounting_pdf(report: dict[str, Any]) -> io.BytesIO:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError as exc:
        raise RuntimeError("reportlab belum terinstall di environment ini") from exc

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    story = [
        Paragraph("Laporan Keuangan SIJI Bintaro", styles["Title"]),
        Paragraph(report["summary"]["periode_label"], styles["Heading3"]),
        Spacer(1, 10),
    ]

    summary_table = Table(
        [
            ["Total Pemasukan", _rp(report["summary"]["total_pemasukan"])],
            ["Total Pengeluaran", _rp(report["summary"]["total_pengeluaran"])],
            ["Net Profit", _rp(report["summary"]["net_profit"])],
            ["Mutasi Belum Ditautkan", str(report["summary"]["unlinked_mutasi"])],
        ],
        colWidths=[7 * cm, 7 * cm],
    )
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9E2F3")),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#A6A6A6")),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ]
        )
    )
    story.extend([summary_table, Spacer(1, 12)])

    if report["summary"]["breakdown_pengeluaran"]:
        story.append(Paragraph("Breakdown Pengeluaran", styles["Heading3"]))
        breakdown = [["Kategori", "Nominal"]]
        for item in report["summary"]["breakdown_pengeluaran"]:
            breakdown.append([item["label"], _rp(item["nominal"])])
        table = Table(breakdown, colWidths=[9 * cm, 5 * cm])
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9E2F3")),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#A6A6A6")),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("ALIGN", (1, 1), (1, -1), "RIGHT"),
                ]
            )
        )
        story.extend([table, Spacer(1, 12)])

    if report["pengeluaran"]:
        story.append(Paragraph("Pengeluaran", styles["Heading3"]))
        top_pengeluaran = [["Tanggal", "Deskripsi", "Kategori", "Nominal"]]
        for item in report["pengeluaran"][:15]:
            top_pengeluaran.append(
                [
                    str(item["tanggal"]),
                    item["deskripsi"][:50],
                    item["kategori_label"][:28],
                    _rp(item["nominal"]),
                ]
            )
        table = Table(top_pengeluaran, colWidths=[2.5 * cm, 7.5 * cm, 4.5 * cm, 3.5 * cm])
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9E2F3")),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#A6A6A6")),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("ALIGN", (-1, 1), (-1, -1), "RIGHT"),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        story.append(table)

    doc.build(story)
    buffer.seek(0)
    return buffer

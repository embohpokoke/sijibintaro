from __future__ import annotations

import csv
import io
import os
import re
from calendar import monthrange
from datetime import date, datetime
from typing import Any, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from pydantic import BaseModel, Field

from auth_api import COOKIE_NAME, _decode_token
from database import get_db_dict
from export_service import build_accounting_excel, build_accounting_pdf
from product_mapping import classify_product

router = APIRouter(prefix="/api/accounting", tags=["accounting"])

MEDIA_STRUK_DIR = "/var/www/sijibintaro/media/struk"


def require_session(request: Request) -> dict[str, Any]:
    token = request.cookies.get(COOKIE_NAME)
    if not token:
        raise HTTPException(status_code=401, detail="Session required")
    try:
        return _decode_token(token)
    except Exception as exc:
        raise HTTPException(status_code=401, detail="Invalid session") from exc


class KategoriCreate(BaseModel):
    nama: str = Field(min_length=1, max_length=100)
    parent_id: Optional[int] = None
    icon: str = Field(default="📦", min_length=1, max_length=10)
    urutan: int = 0


class SupplierCreate(BaseModel):
    nama: str = Field(min_length=1, max_length=200)
    tipe: Optional[str] = Field(default="lainnya")
    kontak: Optional[str] = Field(default=None, max_length=100)
    catatan: Optional[str] = None


class PengeluaranCreate(BaseModel):
    tanggal: Optional[date] = None
    nominal: int = Field(gt=0)
    kategori_id: Optional[int] = None
    supplier_id: Optional[int] = None
    deskripsi: str = Field(min_length=1)
    metode_bayar: str = Field(default="transfer_bca")
    no_referensi: Optional[str] = None
    dicatat_oleh: Optional[str] = None
    sumber_input: str = Field(default="wa_chat")
    foto_struk: Optional[str] = None
    catatan: Optional[str] = None


class PengeluaranUpdate(BaseModel):
    tanggal: Optional[date] = None
    nominal: Optional[int] = Field(default=None, gt=0)
    kategori_id: Optional[int] = None
    supplier_id: Optional[int] = None
    deskripsi: Optional[str] = None
    metode_bayar: Optional[str] = None
    no_referensi: Optional[str] = None
    dicatat_oleh: Optional[str] = None
    sumber_input: Optional[str] = None
    foto_struk: Optional[str] = None
    catatan: Optional[str] = None


class PemasukanManualCreate(BaseModel):
    tanggal: Optional[date] = None
    nama_customer: Optional[str] = None
    layanan: Optional[str] = None
    kategori: Optional[str] = None
    nominal: int = Field(gt=0)
    metode_bayar: str = Field(default="cash")
    dicatat_oleh: Optional[str] = None
    transaction_id: Optional[int] = None
    foto_bukti: Optional[str] = None
    status: str = Field(default="unverified")
    catatan: Optional[str] = None


class MutasiLinkRequest(BaseModel):
    pengeluaran_id: int


class MutasiReconRequest(BaseModel):
    wa_media_path: Optional[str] = None
    recon_notes: Optional[str] = None


def _normalize_tipe_supplier(value: Optional[str]) -> str:
    allowed = {"online", "offline", "vendor", "individu", "lainnya"}
    return value if value in allowed else "lainnya"


def _normalize_metode_pengeluaran(value: Optional[str]) -> str:
    allowed = {"transfer_bca", "cash", "qris", "edc", "lainnya"}
    return value if value in allowed else "transfer_bca"


def _normalize_metode_pemasukan(value: Optional[str]) -> str:
    allowed = {"cash", "transfer", "qris", "edc"}
    return value if value in allowed else "cash"


def _normalize_status_pemasukan(value: Optional[str]) -> str:
    allowed = {"unverified", "verified", "duplicate"}
    return value if value in allowed else "unverified"


def _month_range(bulan: int, tahun: int) -> tuple[date, date]:
    if bulan < 1 or bulan > 12:
        raise HTTPException(status_code=400, detail="bulan harus 1-12")
    if tahun < 2000 or tahun > 2100:
        raise HTTPException(status_code=400, detail="tahun tidak valid")
    last_day = monthrange(tahun, bulan)[1]
    return date(tahun, bulan, 1), date(tahun, bulan, last_day)


def _ensure_category_exists(conn, kategori_id: Optional[int]) -> None:
    if kategori_id is None:
        return
    cur = conn.cursor()
    cur.execute(
        "SELECT id FROM pengeluaran_kategori WHERE id = %s AND aktif = true",
        (kategori_id,),
    )
    if not cur.fetchone():
        raise HTTPException(status_code=404, detail="Kategori tidak ditemukan")


def _ensure_supplier_exists(conn, supplier_id: Optional[int]) -> None:
    if supplier_id is None:
        return
    cur = conn.cursor()
    cur.execute(
        "SELECT id FROM supplier WHERE id = %s AND aktif = true",
        (supplier_id,),
    )
    if not cur.fetchone():
        raise HTTPException(status_code=404, detail="Supplier tidak ditemukan")


def _infer_kategori_layanan(layanan: Optional[str]) -> Optional[str]:
    if not layanan:
        return None
    try:
        return classify_product(layanan).get("kategori")
    except Exception:
        return None


def _parse_money(raw: Any) -> int:
    if raw is None:
        return 0
    if isinstance(raw, (int, float)):
        return int(abs(raw))
    text = str(raw).strip()
    if not text:
        return 0
    cleaned = re.sub(r"[^\d-]", "", text)
    if not cleaned or cleaned == "-":
        return 0
    return abs(int(cleaned))


def _parse_date_value(raw: Any) -> date:
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    text = str(raw or "").strip()
    if not text:
        raise ValueError("tanggal kosong")
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"tanggal tidak dikenali: {text}")


def _normalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")


def _safe_int(value: Any) -> Optional[int]:
    if value in (None, ""):
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _serialize_category_row(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": row["id"],
        "nama": row["nama"],
        "parent_id": row["parent_id"],
        "icon": row["icon"],
        "aktif": row["aktif"],
        "urutan": row["urutan"],
    }


def _get_kategori_flat(conn) -> list[dict[str, Any]]:
    cur = conn.cursor()
    cur.execute(
        """
        SELECT id, nama, parent_id, icon, aktif, urutan
        FROM pengeluaran_kategori
        WHERE aktif = true
        ORDER BY COALESCE(parent_id, id), parent_id NULLS FIRST, urutan, nama
        """
    )
    return [_serialize_category_row(row) for row in cur.fetchall()]


def _get_pengeluaran_detail(conn, pengeluaran_id: int) -> dict[str, Any]:
    cur = conn.cursor()
    cur.execute(
        """
        SELECT
            p.id,
            p.tanggal,
            p.nominal,
            p.kategori_id,
            p.supplier_id,
            p.deskripsi,
            p.metode_bayar,
            p.no_referensi,
            p.dicatat_oleh,
            p.sumber_input,
            p.foto_struk,
            p.catatan,
            p.created_at,
            p.updated_at,
            k.nama AS kategori_nama,
            parent.nama AS kategori_parent_nama,
            parent.icon AS kategori_parent_icon,
            s.nama AS supplier_nama
        FROM pengeluaran p
        LEFT JOIN pengeluaran_kategori k ON k.id = p.kategori_id
        LEFT JOIN pengeluaran_kategori parent ON parent.id = k.parent_id
        LEFT JOIN supplier s ON s.id = p.supplier_id
        WHERE p.id = %s AND p.is_deleted = false
        """,
        (pengeluaran_id,),
    )
    row = cur.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Pengeluaran tidak ditemukan")
    row["kategori_label"] = " > ".join(
        [part for part in [row.get("kategori_parent_nama"), row.get("kategori_nama")] if part]
    ) or "-"
    return row


def _get_pengeluaran_rows(
    conn,
    bulan: Optional[int] = None,
    tahun: Optional[int] = None,
    kategori_id: Optional[int] = None,
    supplier_id: Optional[int] = None,
    tanggal: Optional[date] = None,
    limit: Optional[int] = None,
    offset: Optional[int] = None,
) -> list[dict[str, Any]]:
    where_parts = ["p.is_deleted = false"]
    params: list[Any] = []
    if tanggal:
        where_parts.append("p.tanggal = %s")
        params.append(tanggal)
    elif bulan and tahun:
        start_date, end_date = _month_range(bulan, tahun)
        where_parts.extend(["p.tanggal >= %s", "p.tanggal <= %s"])
        params.extend([start_date, end_date])
    if kategori_id:
        where_parts.append("p.kategori_id = %s")
        params.append(kategori_id)
    if supplier_id:
        where_parts.append("p.supplier_id = %s")
        params.append(supplier_id)
    sql = f"""
        SELECT
            p.id,
            p.tanggal,
            p.nominal,
            p.kategori_id,
            p.supplier_id,
            p.deskripsi,
            p.metode_bayar,
            p.no_referensi,
            p.dicatat_oleh,
            p.sumber_input,
            p.foto_struk,
            p.catatan,
            k.nama AS kategori_nama,
            parent.nama AS kategori_parent_nama,
            s.nama AS supplier_nama
        FROM pengeluaran p
        LEFT JOIN pengeluaran_kategori k ON k.id = p.kategori_id
        LEFT JOIN pengeluaran_kategori parent ON parent.id = k.parent_id
        LEFT JOIN supplier s ON s.id = p.supplier_id
        WHERE {' AND '.join(where_parts)}
        ORDER BY p.tanggal DESC, p.id DESC
    """
    if limit is not None:
        sql += " LIMIT %s"
        params.append(limit)
    if offset is not None:
        sql += " OFFSET %s"
        params.append(offset)
    cur = conn.cursor()
    cur.execute(sql, tuple(params))
    rows = cur.fetchall()
    for row in rows:
        row["kategori_label"] = " > ".join(
            [part for part in [row.get("kategori_parent_nama"), row.get("kategori_nama")] if part]
        ) or "-"
    return rows


def _get_pemasukan_rows(
    conn,
    bulan: Optional[int] = None,
    tahun: Optional[int] = None,
    status: Optional[str] = None,
) -> list[dict[str, Any]]:
    where_parts = ["1 = 1"]
    params: list[Any] = []
    if bulan and tahun:
        start_date, end_date = _month_range(bulan, tahun)
        where_parts.extend(["tanggal >= %s", "tanggal <= %s"])
        params.extend([start_date, end_date])
    if status:
        where_parts.append("status = %s")
        params.append(status)
    cur = conn.cursor()
    cur.execute(
        f"""
        SELECT
            id,
            tanggal,
            nama_customer,
            layanan,
            kategori,
            nominal,
            metode_bayar,
            dicatat_oleh,
            transaction_id,
            foto_bukti,
            status,
            catatan
        FROM pemasukan_manual
        WHERE {' AND '.join(where_parts)}
        ORDER BY tanggal DESC, id DESC
        """,
        tuple(params),
    )
    return cur.fetchall()


def _get_mutasi_rows(conn, bulan: Optional[int] = None, tahun: Optional[int] = None, status_link: Optional[str] = None) -> list[dict[str, Any]]:
    where_parts = ["1 = 1"]
    params: list[Any] = []
    if bulan and tahun:
        start_date, end_date = _month_range(bulan, tahun)
        where_parts.extend(["tanggal >= %s", "tanggal <= %s"])
        params.extend([start_date, end_date])
    if status_link:
        where_parts.append("status_link = %s")
        params.append(status_link)
    cur = conn.cursor()
    cur.execute(
        f"""
        SELECT
            id,
            tanggal,
            no_urut,
            nominal,
            tipe,
            keterangan,
            penerima,
            no_rek_penerima,
            pengeluaran_id,
            status_link,
            wa_media_path,
            recon_notes,
            recon_at
        FROM mutasi_rekening
        WHERE {' AND '.join(where_parts)}
        ORDER BY tanggal DESC, id DESC
        """,
        tuple(params),
    )
    return cur.fetchall()


def _build_summary(conn, bulan: int, tahun: int) -> dict[str, Any]:
    start_date, end_date = _month_range(bulan, tahun)
    cur = conn.cursor()

    cur.execute(
        """
        SELECT COALESCE(SUM(total_tagihan), 0) AS total
        FROM transactions
        WHERE date_of_transaction >= %s
          AND date_of_transaction <= %s
          AND total_tagihan > 0
        """,
        (start_date, end_date),
    )
    total_transactions = int(cur.fetchone()["total"] or 0)

    cur.execute(
        """
        SELECT COALESCE(SUM(nominal), 0) AS total
        FROM pemasukan_manual
        WHERE tanggal >= %s
          AND tanggal <= %s
          AND status != 'duplicate'
          AND transaction_id IS NULL
        """,
        (start_date, end_date),
    )
    total_manual = int(cur.fetchone()["total"] or 0)

    cur.execute(
        """
        SELECT COALESCE(SUM(nominal), 0) AS total
        FROM pengeluaran
        WHERE tanggal >= %s
          AND tanggal <= %s
          AND is_deleted = false
        """,
        (start_date, end_date),
    )
    total_pengeluaran = int(cur.fetchone()["total"] or 0)

    cur.execute(
        """
        SELECT
            COALESCE(parent.nama, child.nama, 'Tanpa Kategori') AS label,
            COALESCE(parent.icon, child.icon, '📦') AS icon,
            COALESCE(SUM(p.nominal), 0) AS nominal
        FROM pengeluaran p
        LEFT JOIN pengeluaran_kategori child ON child.id = p.kategori_id
        LEFT JOIN pengeluaran_kategori parent ON parent.id = child.parent_id
        WHERE p.tanggal >= %s
          AND p.tanggal <= %s
          AND p.is_deleted = false
        GROUP BY 1, 2
        ORDER BY nominal DESC, label
        """,
        (start_date, end_date),
    )
    breakdown_pengeluaran = cur.fetchall()

    cur.execute(
        """
        SELECT label, SUM(nominal) AS nominal
        FROM (
            SELECT COALESCE(group_layanan, nama_layanan, 'Lainnya') AS label,
                   total_tagihan AS nominal
            FROM transactions
            WHERE date_of_transaction >= %s
              AND date_of_transaction <= %s
              AND total_tagihan > 0
            UNION ALL
            SELECT COALESCE(kategori, layanan, 'Manual') AS label,
                   nominal
            FROM pemasukan_manual
            WHERE tanggal >= %s
              AND tanggal <= %s
              AND status != 'duplicate'
              AND transaction_id IS NULL
        ) sumber
        GROUP BY label
        ORDER BY nominal DESC, label
        """,
        (start_date, end_date, start_date, end_date),
    )
    breakdown_pemasukan = cur.fetchall()

    cur.execute(
        """
        SELECT COUNT(*) AS total
        FROM mutasi_rekening
        WHERE tanggal >= %s
          AND tanggal <= %s
          AND status_link = 'unlinked'
        """,
        (start_date, end_date),
    )
    unlinked_mutasi = int(cur.fetchone()["total"] or 0)

    return {
        "periode": {"bulan": bulan, "tahun": tahun, "start": start_date, "end": end_date},
        "periode_label": start_date.strftime("%B %Y"),
        "total_pemasukan_transactions": total_transactions,
        "total_pemasukan_manual": total_manual,
        "total_pemasukan": total_transactions + total_manual,
        "total_pengeluaran": total_pengeluaran,
        "net_profit": (total_transactions + total_manual) - total_pengeluaran,
        "breakdown_pengeluaran": breakdown_pengeluaran,
        "breakdown_pemasukan": breakdown_pemasukan,
        "unlinked_mutasi": unlinked_mutasi,
    }


def _build_report_payload(conn, bulan: int, tahun: int) -> dict[str, Any]:
    summary = _build_summary(conn, bulan, tahun)
    pengeluaran = _get_pengeluaran_rows(conn, bulan=bulan, tahun=tahun)
    pemasukan = _get_pemasukan_rows(conn, bulan=bulan, tahun=tahun)
    mutasi = _get_mutasi_rows(conn, bulan=bulan, tahun=tahun)
    return {
        "summary": summary,
        "pengeluaran": pengeluaran,
        "pemasukan": pemasukan,
        "mutasi": mutasi,
    }


@router.get("/kategori")
async def list_kategori(
    include_children: bool = True,
    _: dict[str, Any] = Depends(require_session),
):
    with get_db_dict() as conn:
        rows = _get_kategori_flat(conn)
    if not include_children:
        return rows
    parent_map: dict[int, dict[str, Any]] = {}
    tree: list[dict[str, Any]] = []
    for row in rows:
        item = {**row, "children": []}
        parent_map[item["id"]] = item
        if item["parent_id"] is None:
            tree.append(item)
    for row in rows:
        if row["parent_id"] is not None and row["parent_id"] in parent_map:
            parent_map[row["parent_id"]]["children"].append(parent_map[row["id"]])
    return tree


@router.post("/kategori")
async def tambah_kategori(
    data: KategoriCreate,
    _: dict[str, Any] = Depends(require_session),
):
    with get_db_dict() as conn:
        if data.parent_id is not None:
            _ensure_category_exists(conn, data.parent_id)
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO pengeluaran_kategori (nama, parent_id, icon, urutan)
            VALUES (%s, %s, %s, %s)
            RETURNING id
            """,
            (data.nama.strip(), data.parent_id, data.icon, data.urutan),
        )
        new_id = cur.fetchone()["id"]
        cur.execute(
            """
            SELECT id, nama, parent_id, icon, aktif, urutan
            FROM pengeluaran_kategori
            WHERE id = %s
            """,
            (new_id,),
        )
        return cur.fetchone()


@router.get("/supplier")
async def list_supplier(
    q: Optional[str] = None,
    _: dict[str, Any] = Depends(require_session),
):
    params: list[Any] = []
    where = "aktif = true"
    if q:
        where += " AND nama ILIKE %s"
        params.append(f"%{q.strip()}%")
    with get_db_dict() as conn:
        cur = conn.cursor()
        cur.execute(
            f"""
            SELECT id, nama, tipe, kontak, catatan, aktif, created_at
            FROM supplier
            WHERE {where}
            ORDER BY nama ASC
            """,
            tuple(params),
        )
        return cur.fetchall()


@router.get("/supplier/search")
async def search_supplier(
    q: str,
    _: dict[str, Any] = Depends(require_session),
):
    if not q.strip():
        raise HTTPException(status_code=400, detail="q wajib diisi")
    return await list_supplier(q=q)


@router.post("/supplier")
async def tambah_supplier(
    data: SupplierCreate,
    _: dict[str, Any] = Depends(require_session),
):
    with get_db_dict() as conn:
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO supplier (nama, tipe, kontak, catatan)
            VALUES (%s, %s, %s, %s)
            ON CONFLICT DO NOTHING
            RETURNING id
            """,
            (data.nama.strip(), _normalize_tipe_supplier(data.tipe), data.kontak, data.catatan),
        )
        inserted = cur.fetchone()
        supplier_id = inserted["id"] if inserted else None
        if supplier_id is None:
            cur.execute(
                "SELECT id FROM supplier WHERE LOWER(nama) = LOWER(%s)",
                (data.nama.strip(),),
            )
            existing = cur.fetchone()
            if not existing:
                raise HTTPException(status_code=500, detail="Gagal menyimpan supplier")
            supplier_id = existing["id"]
        cur.execute(
            """
            SELECT id, nama, tipe, kontak, catatan, aktif, created_at
            FROM supplier
            WHERE id = %s
            """,
            (supplier_id,),
        )
        return cur.fetchone()


@router.post("/pengeluaran")
async def catat_pengeluaran(
    data: PengeluaranCreate,
    session: dict[str, Any] = Depends(require_session),
):
    with get_db_dict() as conn:
        _ensure_category_exists(conn, data.kategori_id)
        _ensure_supplier_exists(conn, data.supplier_id)
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO pengeluaran (
                tanggal, nominal, kategori_id, supplier_id, deskripsi,
                metode_bayar, no_referensi, dicatat_oleh, sumber_input,
                foto_struk, catatan
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
            """,
            (
                data.tanggal or date.today(),
                data.nominal,
                data.kategori_id,
                data.supplier_id,
                data.deskripsi.strip(),
                _normalize_metode_pengeluaran(data.metode_bayar),
                data.no_referensi,
                data.dicatat_oleh or session.get("sub"),
                data.sumber_input or "wa_chat",
                data.foto_struk,
                data.catatan,
            ),
        )
        new_id = cur.fetchone()["id"]
        return _get_pengeluaran_detail(conn, new_id)


@router.get("/pengeluaran")
async def list_pengeluaran(
    bulan: Optional[int] = None,
    tahun: Optional[int] = None,
    kategori_id: Optional[int] = None,
    supplier_id: Optional[int] = None,
    tanggal: Optional[date] = None,
    limit: int = Query(default=50, ge=1, le=500),
    offset: int = Query(default=0, ge=0),
    _: dict[str, Any] = Depends(require_session),
):
    with get_db_dict() as conn:
        items = _get_pengeluaran_rows(
            conn,
            bulan=bulan,
            tahun=tahun,
            kategori_id=kategori_id,
            supplier_id=supplier_id,
            tanggal=tanggal,
            limit=limit,
            offset=offset,
        )
        return {"items": items, "limit": limit, "offset": offset}


@router.get("/pengeluaran/{pengeluaran_id}")
async def detail_pengeluaran(
    pengeluaran_id: int,
    _: dict[str, Any] = Depends(require_session),
):
    with get_db_dict() as conn:
        return _get_pengeluaran_detail(conn, pengeluaran_id)


@router.patch("/pengeluaran/{pengeluaran_id}")
async def update_pengeluaran(
    pengeluaran_id: int,
    data: PengeluaranUpdate,
    _: dict[str, Any] = Depends(require_session),
):
    payload = data.model_dump(exclude_unset=True)
    if not payload:
        with get_db_dict() as conn:
            return _get_pengeluaran_detail(conn, pengeluaran_id)
    with get_db_dict() as conn:
        _get_pengeluaran_detail(conn, pengeluaran_id)
        if "kategori_id" in payload:
            _ensure_category_exists(conn, payload["kategori_id"])
        if "supplier_id" in payload:
            _ensure_supplier_exists(conn, payload["supplier_id"])
        if "metode_bayar" in payload:
            payload["metode_bayar"] = _normalize_metode_pengeluaran(payload["metode_bayar"])
        if "deskripsi" in payload and payload["deskripsi"] is not None:
            payload["deskripsi"] = payload["deskripsi"].strip()
        set_parts = []
        params: list[Any] = []
        for key, value in payload.items():
            set_parts.append(f"{key} = %s")
            params.append(value)
        set_parts.append("updated_at = NOW()")
        params.append(pengeluaran_id)
        cur = conn.cursor()
        cur.execute(
            f"""
            UPDATE pengeluaran
            SET {', '.join(set_parts)}
            WHERE id = %s
            """,
            tuple(params),
        )
        return _get_pengeluaran_detail(conn, pengeluaran_id)


@router.delete("/pengeluaran/{pengeluaran_id}")
async def delete_pengeluaran(
    pengeluaran_id: int,
    _: dict[str, Any] = Depends(require_session),
):
    with get_db_dict() as conn:
        _get_pengeluaran_detail(conn, pengeluaran_id)
        cur = conn.cursor()
        cur.execute(
            """
            UPDATE pengeluaran
            SET is_deleted = true, updated_at = NOW()
            WHERE id = %s
            """,
            (pengeluaran_id,),
        )
        return {"ok": True, "id": pengeluaran_id}


@router.post("/pemasukan")
async def catat_pemasukan_manual(
    data: PemasukanManualCreate,
    session: dict[str, Any] = Depends(require_session),
):
    kategori = data.kategori or _infer_kategori_layanan(data.layanan)
    with get_db_dict() as conn:
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO pemasukan_manual (
                tanggal, nama_customer, layanan, kategori, nominal,
                metode_bayar, dicatat_oleh, transaction_id, foto_bukti,
                status, catatan
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
            """,
            (
                data.tanggal or date.today(),
                data.nama_customer,
                data.layanan,
                kategori,
                data.nominal,
                _normalize_metode_pemasukan(data.metode_bayar),
                data.dicatat_oleh or session.get("sub"),
                data.transaction_id,
                data.foto_bukti,
                _normalize_status_pemasukan(data.status),
                data.catatan,
            ),
        )
        new_id = cur.fetchone()["id"]
        cur.execute(
            """
            SELECT
                id, tanggal, nama_customer, layanan, kategori, nominal,
                metode_bayar, dicatat_oleh, transaction_id, foto_bukti,
                status, catatan
            FROM pemasukan_manual
            WHERE id = %s
            """,
            (new_id,),
        )
        return cur.fetchone()


@router.get("/pemasukan")
async def list_pemasukan_manual(
    bulan: Optional[int] = None,
    tahun: Optional[int] = None,
    status: Optional[str] = None,
    _: dict[str, Any] = Depends(require_session),
):
    with get_db_dict() as conn:
        return {"items": _get_pemasukan_rows(conn, bulan=bulan, tahun=tahun, status=status)}


@router.post("/mutasi/import")
async def import_mutasi(
    file: UploadFile = File(...),
    _: dict[str, Any] = Depends(require_session),
):
    ext = os.path.splitext(file.filename or "")[1].lower()
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="File kosong")

    rows: list[dict[str, Any]] = []
    if ext == ".csv":
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            rows.append({_normalize_header(k): v for k, v in row.items()})
    elif ext in {".xlsx", ".xlsm"}:
        workbook = load_workbook(io.BytesIO(content), data_only=True)
        sheet = workbook.active
        header_row = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        headers = [_normalize_header(value) for value in header_row]
        for values in sheet.iter_rows(min_row=2, values_only=True):
            if not any(values):
                continue
            rows.append({headers[idx]: value for idx, value in enumerate(values)})
    else:
        raise HTTPException(status_code=400, detail="Format file harus CSV atau XLSX")

    parsed_rows: list[tuple[Any, ...]] = []
    for row in rows:
        tanggal_raw = row.get("tanggal") or row.get("trx_date") or row.get("date")
        nominal_raw = row.get("nominal") or row.get("amount") or row.get("jumlah")
        tipe_raw = str(row.get("tipe") or row.get("type") or "").strip().lower()
        no_urut = row.get("no_urut") or row.get("no") or row.get("urut")
        keterangan = row.get("keterangan") or row.get("description") or row.get("ket")
        penerima = row.get("penerima") or row.get("receiver")
        no_rek = row.get("no_rek_penerima") or row.get("rekening") or row.get("account_number")
        try:
            tanggal = _parse_date_value(tanggal_raw)
        except ValueError:
            continue
        nominal = _parse_money(nominal_raw)
        if nominal <= 0:
            continue
        if tipe_raw not in {"debit", "kredit"}:
            tipe_raw = "debit" if "-" in str(nominal_raw) else "kredit"
        parsed_rows.append((tanggal, no_urut, nominal, tipe_raw, keterangan, penerima, no_rek))

    if not parsed_rows:
        raise HTTPException(status_code=400, detail="Tidak ada baris mutasi yang valid")

    inserted = 0
    with get_db_dict() as conn:
        cur = conn.cursor()
        for item in parsed_rows:
            cur.execute(
                """
                INSERT INTO mutasi_rekening (
                    tanggal, no_urut, nominal, tipe, keterangan, penerima, no_rek_penerima
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                RETURNING id
                """,
                item,
            )
            if cur.fetchone():
                inserted += 1
    return {"ok": True, "inserted": inserted}


@router.get("/mutasi")
async def list_mutasi(
    bulan: Optional[int] = None,
    tahun: Optional[int] = None,
    status_link: Optional[str] = "unlinked",
    _: dict[str, Any] = Depends(require_session),
):
    with get_db_dict() as conn:
        return {"items": _get_mutasi_rows(conn, bulan=bulan, tahun=tahun, status_link=status_link)}


@router.patch("/mutasi/{mutasi_id}/link")
async def link_mutasi_ke_pengeluaran(
    mutasi_id: int,
    data: MutasiLinkRequest,
    _: dict[str, Any] = Depends(require_session),
):
    with get_db_dict() as conn:
        _get_pengeluaran_detail(conn, data.pengeluaran_id)
        cur = conn.cursor()
        cur.execute("SELECT id FROM mutasi_rekening WHERE id = %s", (mutasi_id,))
        if not cur.fetchone():
            raise HTTPException(status_code=404, detail="Mutasi tidak ditemukan")
        cur.execute(
            """
            UPDATE mutasi_rekening
            SET pengeluaran_id = %s, status_link = 'linked'
            WHERE id = %s
            """,
            (data.pengeluaran_id, mutasi_id),
        )
        return {"ok": True, "id": mutasi_id, "pengeluaran_id": data.pengeluaran_id}


@router.get("/mutasi/unlinked-incoming")
async def list_unlinked_incoming_mutasi(
    days: int = Query(default=14, ge=1, le=90),
    _: dict[str, Any] = Depends(require_session),
):
    """List unlinked incoming (kredit) mutasi — customer payments that need WA proof reconciliation."""
    with get_db_dict() as conn:
        since = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d")
        cur = conn.cursor()
        cur.execute(
            """
            SELECT
                id, tanggal, no_urut, nominal, tipe,
                keterangan, penerima, no_rek_penerima,
                pengeluaran_id, status_link,
                wa_media_path, recon_notes, recon_at
            FROM mutasi_rekening
            WHERE tipe = 'kredit'
              AND status_link = 'unlinked'
              AND tanggal >= %s
            ORDER BY tanggal DESC, id DESC
            """,
            (since,),
        )
        rows = cur.fetchall()
        cols = [d[0] for d in cur.description]
        return {"items": [dict(zip(cols, row)) for row in rows], "count": len(rows), "days": days}


@router.get("/mutasi/{mutasi_id}")
async def get_mutasi_detail(
    mutasi_id: int,
    _: dict[str, Any] = Depends(require_session),
):
    with get_db_dict() as conn:
        cur = conn.cursor()
        cur.execute(
            """
            SELECT
                id, tanggal, no_urut, nominal, tipe,
                keterangan, penerima, no_rek_penerima,
                pengeluaran_id, status_link,
                wa_media_path, recon_notes, recon_at
            FROM mutasi_rekening
            WHERE id = %s
            """,
            (mutasi_id,),
        )
        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Mutasi tidak ditemukan")
        cols = [d[0] for d in cur.description]
        return dict(zip(cols, row))


@router.patch("/mutasi/{mutasi_id}/recon")
async def recon_mutasi_with_wa(
    mutasi_id: int,
    data: MutasiReconRequest,
    _: dict[str, Any] = Depends(require_session),
):
    """Reconcile a mutasi (incoming transfer) with a WA payment proof image.
    If wa_media_path is null/empty, unlink the mutasi (set status back to unlinked)."""
    with get_db_dict() as conn:
        cur = conn.cursor()
        cur.execute("SELECT id FROM mutasi_rekening WHERE id = %s", (mutasi_id,))
        if not cur.fetchone():
            raise HTTPException(status_code=404, detail="Mutasi tidak ditemukan")
        is_linking = bool(data.wa_media_path)
        cur.execute(
            """
            UPDATE mutasi_rekening
            SET wa_media_path = %s,
                recon_notes = %s,
                recon_at = CASE WHEN %s THEN NOW() ELSE NULL END,
                status_link = CASE WHEN %s THEN 'linked' ELSE 'unlinked' END
            WHERE id = %s
            """,
            (data.wa_media_path, data.recon_notes, is_linking, is_linking, mutasi_id),
        )
        return {"ok": True, "id": mutasi_id, "wa_media_path": data.wa_media_path, "status": "linked" if is_linking else "unlinked"}


@router.get("/laporan/summary")
async def summary_bulanan(
    bulan: int,
    tahun: int,
    _: dict[str, Any] = Depends(require_session),
):
    with get_db_dict() as conn:
        return _build_summary(conn, bulan, tahun)


@router.get("/laporan/export")
async def export_laporan(
    bulan: int,
    tahun: int,
    format: str = Query(default="xlsx", pattern="^(xlsx|pdf)$"),
    _: dict[str, Any] = Depends(require_session),
):
    with get_db_dict() as conn:
        report = _build_report_payload(conn, bulan, tahun)

    filename_base = f"laporan-keuangan-siji-{tahun}-{bulan:02d}"
    if format == "xlsx":
        buffer = build_accounting_excel(report)
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename_base}.xlsx"'},
        )

    buffer = build_accounting_pdf(report)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename_base}.pdf"'},
    )
